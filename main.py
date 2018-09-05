from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl import styles
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.worksheet.cell_range import CellRange
from datetime import date
from datetime import timedelta


def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days)):
        yield start_date + timedelta(n)


def set_border(ws, cell_range, border_style):
    rows = ws[cell_range]
    side = Side(border_style=border_style , color='FF000000')

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def main():
    start_date = date(2016, 1, 1)
    end_date = date(2018, 12, 31)
    date_list = daterange(start_date, end_date)
    wb = Workbook()
    dest_filename = 'empty_book.xlsx'
    ws = wb.active
    ws.title = 'ah ha'
    holydays_fill = PatternFill('solid', fgColor='CCCCCC')
    column_width = 3
    colored_rows = 10
    col = 2
    row = 1
    for this_date in date_list:
        ws.column_dimensions[get_column_letter(col)].width = column_width
        year_cell = ws.cell(column=col, row=row, value=int(this_date.strftime('%Y')))
        year_cell.font = styles.Font(bold=True)
        year_cell.alignment = Alignment(horizontal='center')
        month_cell = ws.cell(column=col, row=row+1, value=this_date.strftime('%B'))
        month_cell.font = styles.Font(bold=True)
        month_cell.alignment = Alignment(horizontal='center')
        day_cell = ws.cell(column=col, row=row+2, value=int(this_date.strftime('%d')))
        day_cell.font = styles.Font(bold=True)
        day_cell.alignment = Alignment(horizontal='center')
        if this_date.weekday() >= 5:
            for i in range(day_cell.row, day_cell.row+colored_rows):
                ws.cell(column=col, row=i).fill = holydays_fill
        col = col + 1

    ranges = []
    # calculate year ranges
    this_year = ws.cell(row=1, column=2)
    for i in range(2, col+1):
        if ws.cell(row=1, column=i).value != this_year.value:
            ranges.append({'type': 'year', 'range': [1, 1, column_index_from_string(this_year.column), i-1]})
            this_year = ws.cell(row=1, column=i)
        if this_year.value is None:
            break

    # calculate month ranges
    this_year = ws.cell(row=2, column=2)
    for i in range(2, col + 1):
        if ws.cell(row=2, column=i).value != this_year.value:
            ranges.append({'type': 'month', 'range': [2, 2, column_index_from_string(this_year.column), i - 1]})
            this_year = ws.cell(row=2, column=i)
        if this_year.value is None:
            break

    # merge ranges and add borders
    for this_range_dict in ranges:
        this_range = this_range_dict['range']
        ws.merge_cells(start_row=this_range[0], end_row=this_range[1], start_column=this_range[2], end_column=this_range[3])
        if this_range_dict['type'] == 'year':
            set_border(ws, CellRange(min_row=this_range[0], max_row=this_range[1],
                                     min_col=this_range[2], max_col=this_range[3]).coord, 'thick')
        if this_range_dict['type'] == 'month':
            # month thin line
            set_border(ws, CellRange(min_row=this_range[0], max_row=this_range[1],
                                     min_col=this_range[2], max_col=this_range[3]).coord, 'thick')
            # day thin line
            set_border(ws, CellRange(min_row=this_range[0]+1, max_row=this_range[1]+1,
                                     min_col=this_range[2], max_col=this_range[3]).coord, 'thin')
            # month thick line
            set_border(ws, CellRange(min_row=this_range[0], max_row=this_range[1]+colored_rows,
                                     min_col=this_range[2], max_col=this_range[3]).coord, 'thick')
    wb.save(filename=dest_filename)


if __name__ == "__main__":
    main()
